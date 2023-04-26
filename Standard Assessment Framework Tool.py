from openpyxl import workbook,load_workbook

#Function to turn skill keys to grade percentages
def Skl2Grd(args):
    sklgrdkey={4:1,3.5:.9,3:.8,2.5:.75,2:.6,1.5:.5,1:.45,.5:.4,0:.35}
    grds=[]
    for arg in args:
        for val in arg:
            if type(val)==int:
                grds.append(sklgrdkey[val])
            else:
                grds.append('exc')
    return grds

#Taking percentages and porportioning them, leaving any strings as-is (for excused, absent, etc.)
def PortionGrd(args):
    finalgrades=[(val*total if type(val)==int or type(val)==float else val) for val in args]
    return finalgrades

#Taking excel file input, loading workbook and setting active workpages with openpyxl
name=input('Please copy and paste the name of your excel file (including extension) here.')
wb=load_workbook(filename=name)
ws=wb.active

#Taking user input to find min/max row and column to iterate over/extract values from.
x=input('input the cell where the skills column begins (i.e."G2")')
m_col=int(ord(x[0].upper())-64)
mn_row=int(x[1:])
y=input('input the cell of the last skill in the assignment.')
mx_row=int(y[1:])

#Iterating over column section specified, creating list of skill values, and running them through skill-to-grade function
sklist=[]
for skill in ws.iter_cols(m_col,m_col,mn_row,mx_row,values_only=True):
   sklist.append(skill)
percents=Skl2Grd(sklist)


total=float(input("Out of how many points?"))
finalgrades=PortionGrd(percents)

#Writing a new excel worksheet on doc with final grades
ws2=wb.create_sheet(f"Output Grades out of {total}")

for row, val in enumerate(finalgrades,start=1):
    ws2.cell(row=row,column=1).value=val
wb.save(name)

print("Done!  Please check your document.") 